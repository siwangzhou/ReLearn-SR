import os.path
import logging
import time
import argparse

import lpips
import torch
from torch import optim
from tqdm import tqdm
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import options.options as option
import utils.util as util
from data import create_dataset, create_dataloader
from models import create_model
import torch.nn.functional as F

# 训练参数配置
TRAIN_CONFIG = {
    'iterations': 100,
    'learning_rate': 0.001,
    'step_size': 20,
    'gamma': 0.8
}
loss_fn = lpips.LPIPS(net='alex')


opt_yml = 'options/test/test_P2P_HCD_CARN_conv_4X.yml'

def calculate_metrics(y_prime, HR):
    """计算所有评估指标"""
    psnr, ssim, psnr_y, ssim_y = util.calculate_psnr_ssim(y_prime[0].detach().cpu().numpy(),
                                                         HR[0].detach().cpu().numpy())
    lpips_distance = calculate_lpips(y_prime[0].detach().cpu(), HR[0].detach().cpu())
    return {
        'psnr': psnr,
        'ssim': ssim,
        'psnr_y': psnr_y,
        'ssim_y': ssim_y,
        'lpips': lpips_distance
    }

def update_best_metrics(current_metrics, best_metrics, iteration):
    """更新最佳指标"""
    updated = False
    for metric in ['psnr', 'ssim', 'psnr_y', 'ssim_y']:
        if current_metrics[metric] > best_metrics[metric]['value']:
            best_metrics[metric]['value'] = current_metrics[metric]
            best_metrics[metric]['iteration'] = iteration
            updated = True
    # LPIPS越小越好
    if current_metrics['lpips'] < best_metrics['lpips']['value']:
        best_metrics['lpips']['value'] = current_metrics['lpips']
        best_metrics['lpips']['iteration'] = iteration
        updated = True
    return updated

def calculate_lpips(img1_tensor,img2_tensor):
    lpips_distance = loss_fn(img1_tensor, img2_tensor)
    return lpips_distance.item()


parser = argparse.ArgumentParser()
parser.add_argument('--opt', type=str, default=opt_yml)
args = parser.parse_args()
opt = option.parse(args.opt, is_train=False)
opt = option.dict_to_nonedict(opt)
device_id = torch.cuda.current_device()
device = torch.device('cuda:{}'.format(device_id) if torch.cuda.is_available() else 'cpu')
gpu_info = util.get_gpu_info()

# mkdir and logger
util.mkdirs((path for key, path in opt['path'].items() if not key == 'experiments_root'
             and 'pretrain_model' not in key and 'resume' not in key and 'load_submodule' not in key))
util.setup_logger('base', opt['path']['log'], 'test_' + opt['name'], level=logging.INFO,
                  screen=True, tofile=True)
logger = logging.getLogger('base')
util.set_random_seed(0)

# 创建 test dataset and dataloader
test_loaders = []
for phase, dataset_opt in sorted(opt['datasets'].items()):
    test_set = create_dataset(dataset_opt)
    test_loader = create_dataloader(test_set, dataset_opt)
    logger.info('Number of test images in [{:s}]: {:d}'.format(dataset_opt['name'], len(test_set)))
    test_loaders.append(test_loader)

# 创建 model 并冻结网络参数
model = create_model(opt)
model.freeze()

for test_loader in test_loaders:
    test_set_name = test_loader.dataset.opt['name']
    logger.info('\n\nTesting [{:s}]...'.format(test_set_name))
    dataset_dir = os.path.join(opt['path']['results_root'], test_set_name)
    util.mkdir(dataset_dir)

    # 初始化所有指标的平均值
    metrics_avg = {
        'first': {'psnr': 0, 'ssim': 0, 'psnr_y': 0, 'ssim_y': 0, 'lpips': 0},
        'max': {'psnr': 0, 'ssim': 0, 'psnr_y': 0, 'ssim_y': 0, 'lpips': 0},
        'increment': {'psnr': 0, 'ssim': 0, 'psnr_y': 0, 'ssim_y': 0, 'lpips': 0}
    }
    results_data = []

    for test_data in test_loader:

        img_path = test_data['GT_path'][0]
        img_name = os.path.splitext(os.path.basename(img_path))[0]
        output_dir = os.path.join(dataset_dir, img_name)
        util.mkdir(output_dir)

        # 初始化最佳指标
        best_metrics = {
            'psnr': {'value': -float('inf'), 'iteration': 0},
            'ssim': {'value': -float('inf'), 'iteration': 0},
            'psnr_y': {'value': -float('inf'), 'iteration': 0},
            'ssim_y': {'value': -float('inf'), 'iteration': 0},
            'lpips': {'value': float('inf'), 'iteration': 0}
        }
        first_metrics = None
        best_sr = None
        first_sr = None
        best_delta = None
        psnr_max_time = 0  # 新增变量，记录PSNR最佳值时的时间

        HR = test_data['GT'].to(device)

        # 生成LR图像
        with torch.no_grad():
            LR = model.get_downsample(HR)
            LR.requires_grad_(False)

        # 初始化delta
        delta = torch.zeros_like(LR).to(device)
        delta.requires_grad = True

        # 定义损失函数和优化器
        lossMSE = torch.nn.MSELoss(reduction='mean').to(device)
        optimizer = optim.Adam([delta], lr=TRAIN_CONFIG['learning_rate'])
        # scheduler = optim.lr_scheduler.StepLR(optimizer, TRAIN_CONFIG['step_size'], TRAIN_CONFIG['gamma'])
        scheduler = optim.lr_scheduler.CosineAnnealingLR(optimizer, T_max=TRAIN_CONFIG['iterations'])

        # ======================= 生成第一次SR图片代码 ======================
        # SR = model.get_upsample(LR)
        # SR_img = util.tenso2img(SR)
        # util.save_img(SR_img, os.path.join(output_dir, f'{img_name}_first.png'))
        # continue
        # =======================

        start_time = time.time()

        bar = tqdm(range(TRAIN_CONFIG['iterations']))
        for i in bar:
            torch.cuda.empty_cache()
            optimizer.zero_grad()

            LR_new = LR + delta
            y_prime, x_prime_downsample = model.my_test(LR_new)
            
            loss = lossMSE(x_prime_downsample, LR)
            loss.backward()
            
            optimizer.step()
            scheduler.step()
            delta.grad.data.zero_()

            # 计算当前指标
            current_metrics = calculate_metrics(y_prime, HR)
            
            if i == 0:
                first_metrics = current_metrics
                first_sr = y_prime[0].detach().cpu()

            # 更新最佳指标
            if update_best_metrics(current_metrics, best_metrics, i):
                best_sr = y_prime[0].detach().cpu()
                best_delta = delta
                # 记录PSNR最佳值时的时间
                if current_metrics['psnr'] > best_metrics['psnr']['value'] - 1e-4:  # 考虑浮点误差
                    psnr_max_time = time.time() - start_time

            # 更新进度条
            bar.set_postfix_str(
                f"PSNR: {current_metrics['psnr']:.4f}, "
                f"SSIM: {current_metrics['ssim']:.4f}, "
                f"LPIPS: {current_metrics['lpips']:.4f}")
            bar.set_description(f'Processing: {img_name}')

        finish_time = time.time()


        util.save_perturbation(best_delta, os.path.join(output_dir, f'{img_name}_delta_best_view.png'))
        # 将best_delta 进行超分
        best_delta_sr = model.get_upsample(best_delta)
        best_delta = best_delta[0].detach().cpu()

        # 对HR进行下采样得到LR_bic
        hr_tensor = HR[0].detach().cpu()
        lr_bic = F.interpolate(
            hr_tensor.unsqueeze(0),  # 添加批次维度
            scale_factor=0.5,  # 2倍下采样
            mode='bicubic',
            align_corners=False
        ).squeeze(0)  # 移除批次维度
        
        # 将LR_bic转换为图像并保存
        lr_bic_img = util.tensor2img(lr_bic)
        util.save_img(lr_bic_img, os.path.join(output_dir, f'{img_name}_lr_bic.png'))



        # 保存结果
        best_sr_img = util.tensor2img(best_sr)
        first_sr_img = util.tensor2img(first_sr)
        best_delta_img = util.tensor2img(best_delta)
        first_lr_img = util.tensor2img(LR[0].detach().cpu())
        best_delta_sr_img = util.tensor2img(best_delta_sr[0].detach().cpu())
        util.save_img(best_sr_img, os.path.join(output_dir, f'{img_name}_best.png'))
        util.save_img(first_sr_img, os.path.join(output_dir, f'{img_name}_first.png'))
        util.save_img(best_delta_img, os.path.join(output_dir, f'{img_name}_delta_best.png'))
        util.save_img(best_delta_sr_img, os.path.join(output_dir, f'{img_name}_delta_best_sr.png'))
        util.save_img(first_lr_img, os.path.join(output_dir, f'{img_name}_first_lr.png'))
        # 记录结果数据
        result = {
            'Image Name': img_name,
            'PSNR_0': f"{first_metrics['psnr']:.4f}",
            'PSNR_Max': f"{best_metrics['psnr']['value']:.4f}",
            'PSNR_Increment': f"{(best_metrics['psnr']['value'] - first_metrics['psnr']):.4f}",
            'SSIM_0': f"{first_metrics['ssim']:.4f}",
            'SSIM_Max': f"{best_metrics['ssim']['value']:.4f}",
            'SSIM_Increment': f"{(best_metrics['ssim']['value'] - first_metrics['ssim']):.4f}",
            'PSNR-Y_0': f"{first_metrics['psnr_y']:.4f}",
            'PSNR-Y_MAX': f"{best_metrics['psnr_y']['value']:.4f}",
            'PSNR-Y_Increment': f"{(best_metrics['psnr_y']['value'] - first_metrics['psnr_y']):.4f}",
            'SSIM-Y_0': f"{first_metrics['ssim_y']:.4f}",
            'SSIM-Y_MAX': f"{best_metrics['ssim_y']['value']:.4f}",
            'SSIM-Y_Increment': f"{(best_metrics['ssim_y']['value'] - first_metrics['ssim_y']):.4f}",
            'LPIPS_0': f"{first_metrics['lpips']:.4f}",
            'LPIPS_Min': f"{best_metrics['lpips']['value']:.4f}",
            'LPIPS_Decrement': f"{(first_metrics['lpips'] - best_metrics['lpips']['value']):.4f}",
            'Peek At': best_metrics['psnr']['iteration'] + 1,
            'PSNR_Max_Time': f'{psnr_max_time:.2f}s',  # 新增：PSNR最佳值时的时间
            'Used Time': f'{finish_time - start_time:.2f}s'
        }
        results_data.append(result)

        # 更新平均值
        for metric in ['psnr', 'ssim', 'psnr_y', 'ssim_y', 'lpips']:
            metrics_avg['first'][metric] += first_metrics[metric]
            metrics_avg['max'][metric] += best_metrics[metric]['value']
            metrics_avg['increment'][metric] += (best_metrics[metric]['value'] - first_metrics[metric])

        # 记录日志
        logger.info(f"\nCurrent Image: {img_name}")
        for metric in ['psnr', 'ssim', 'psnr_y', 'ssim_y']:
            logger.info(f"{metric.upper()}: First={first_metrics[metric]:.4f}, "
                       f"Max={best_metrics[metric]['value']:.4f}, "
                       f"Increment={(best_metrics[metric]['value'] - first_metrics[metric]):.4f}")
        logger.info(f"LPIPS: First={first_metrics['lpips']:.4f}, "
                   f"Min={best_metrics['lpips']['value']:.4f}, "
                   f"Decrement={(first_metrics['lpips'] - best_metrics['lpips']['value']):.4f}")
        logger.info(f"Time: PSNR_Max_Time={psnr_max_time:.2f}s, Total_Time={finish_time - start_time:.2f}s")

    # 计算所有平均值
    num_images = len(test_loader)
    average_data = {
        'Image Name': 'Average',
        'Model': opt['model'],
        'PSNR_0': f"{metrics_avg['first']['psnr'] / num_images:.5f}",
        'PSNR_Max': f"{metrics_avg['max']['psnr'] / num_images:.5f}",
        'PSNR_Increment': f"{metrics_avg['increment']['psnr'] / num_images:.5f}",
        'SSIM_0': f"{metrics_avg['first']['ssim'] / num_images:.5f}",
        'SSIM_Max': f"{metrics_avg['max']['ssim'] / num_images:.5f}",
        'SSIM_Increment': f"{metrics_avg['increment']['ssim'] / num_images:.5f}",
        'PSNR-Y_0': f"{metrics_avg['first']['psnr_y'] / num_images:.5f}",
        'PSNR-Y_MAX': f"{metrics_avg['max']['psnr_y'] / num_images:.5f}",
        'PSNR-Y_Increment': f"{metrics_avg['increment']['psnr_y'] / num_images:.5f}",
        'SSIM-Y_0': f"{metrics_avg['first']['ssim_y'] / num_images:.5f}",
        'SSIM-Y_MAX': f"{metrics_avg['max']['ssim_y'] / num_images:.5f}",
        'SSIM-Y_Increment': f"{metrics_avg['increment']['ssim_y'] / num_images:.5f}",
        'LPIPS_0': f"{metrics_avg['first']['lpips'] / num_images:.5f}",
        'LPIPS_Min': f"{metrics_avg['max']['lpips'] / num_images:.5f}",
        'LPIPS_Decrement': f"{-metrics_avg['increment']['lpips'] / num_images:.5f}",
        'Peek At': '',
        'PSNR_Max_Time': '',  # 新增：PSNR最佳值时的时间
        'Used Time': ''
    }

    # 创建并保存Excel文件
    excel_path = os.path.join(dataset_dir, f'{test_set_name}_{opt["model"]}_results_{util.get_timestamp()}.xlsx')
    
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        results_df = pd.DataFrame(results_data)
        
        # 确保average_data包含与results_data相同的列
        for col in results_df.columns:
            if col not in average_data and col != 'Model':
                average_data[col] = ''
        
        # 添加平均值行
        results_df = pd.concat([results_df, pd.DataFrame([average_data])], ignore_index=True)
        
        # 添加GPU信息
        gpu_info_row = {col: '' for col in results_df.columns}
        gpu_info_row['Image Name'] = gpu_info
        results_df = pd.concat([pd.DataFrame([gpu_info_row]), results_df], ignore_index=True)
        
        results_df.to_excel(writer, sheet_name='Sheet1', index=False)
        
        # 设置Excel样式
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        # 定义样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # 设置表头格式
        for col in range(1, len(results_df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # 设置数据单元格格式
        for row in range(2, len(results_df) + 1):
            for col in range(1, len(results_df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal='center')

                if row == 2:  # GPU信息行
                    cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                    cell.font = Font(bold=True)
        
        # 设置列宽
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            worksheet.column_dimensions[column].width = max_length + 4

        logger.info(f'Results saved to {excel_path}')
